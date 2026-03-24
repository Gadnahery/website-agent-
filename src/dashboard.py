from __future__ import annotations

from io import BytesIO
import os
import re
import threading
import uuid
import webbrowser
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Callable

from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from cache import (
    get_call_sheet_path,
    get_leads_excel_path,
    get_scraper_results_path,
    load_leads,
    save_leads,
)
from classes.WebsiteSalesAgent import AgentStoppedError, WebsiteSalesAgent
from config import (
    get_build_package_output_dir,
    get_country,
    get_google_drive_folder_id,
    get_minimum_rating,
    get_minimum_review_count,
    get_must_have_phone,
    get_proposal_output_dir,
    get_require_missing_website,
    get_scraper_concurrency,
    get_scraper_depth,
    get_scraper_fast_mode,
    get_scraper_lang,
    get_scraper_radius,
    get_target_cities,
    get_target_niches,
    get_target_queries,
)
from constants import LEAD_STATUSES
from drive_sync import DriveSync
from utils import assert_folder_structure

app = Flask(__name__, template_folder="templates", static_folder="static")
assert_folder_structure()


def _now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _parse_int(value: Any, default: int) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _parse_float(value: Any, default: float) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _parse_bool(value: Any, default: bool) -> bool:
    if isinstance(value, bool):
        return value
    if value in (None, ""):
        return default
    return str(value).strip().lower() in {"1", "true", "yes", "on"}


def _split_terms(value: Any) -> list[str]:
    if isinstance(value, list):
        raw_items = value
    else:
        text = str(value or "")
        raw_items = []
        for line in text.replace(";", "\n").splitlines():
            raw_items.extend(part.strip() for part in line.split(","))

    items: list[str] = []
    for item in raw_items:
        cleaned = str(item).strip()
        if cleaned and cleaned not in items:
            items.append(cleaned)
    return items


def _defaults_snapshot() -> dict[str, Any]:
    return {
        "country": get_country(),
        "target_cities": get_target_cities(),
        "target_niches": get_target_niches(),
        "target_queries": get_target_queries(),
        "must_have_phone": get_must_have_phone(),
        "require_missing_website": get_require_missing_website(),
        "minimum_review_count": get_minimum_review_count(),
        "minimum_rating": get_minimum_rating(),
        "scraper_depth": get_scraper_depth(),
        "scraper_concurrency": get_scraper_concurrency(),
        "scraper_fast_mode": get_scraper_fast_mode(),
        "scraper_radius": get_scraper_radius(),
        "scraper_lang": get_scraper_lang(),
        "google_drive_folder_id": get_google_drive_folder_id(),
    }


def _search_presets() -> list[dict[str, Any]]:
    return [
        {
            "name": "Dar salons",
            "locations": ["Dar es Salaam"],
            "business_types": ["salon", "barbershop", "beauty spa"],
        },
        {
            "name": "Dar halls",
            "locations": ["Dar es Salaam"],
            "business_types": ["event hall", "conference hall", "wedding venue"],
        },
        {
            "name": "Arusha tourism",
            "locations": ["Arusha", "Moshi"],
            "business_types": ["tour operator", "safari lodge", "travel agency"],
        },
        {
            "name": "Neighborhood search",
            "target_queries": [
                "salons in Mikocheni, Dar es Salaam",
                "event halls in Kariakoo, Dar es Salaam",
            ],
        },
    ]


def _build_discovery_profile(payload: dict[str, Any]) -> dict[str, Any]:
    defaults = _defaults_snapshot()
    profile: dict[str, Any] = {}

    country = str(payload.get("country", "")).strip()
    if country:
        profile["country"] = country

    locations = _split_terms(payload.get("locations"))
    if locations:
        profile["target_cities"] = locations

    business_types = _split_terms(payload.get("business_types"))
    if business_types:
        profile["target_niches"] = business_types

    raw_queries = _split_terms(payload.get("target_queries"))
    if raw_queries:
        profile["target_queries"] = raw_queries

    profile["must_have_phone"] = _parse_bool(
        payload.get("must_have_phone"), defaults["must_have_phone"]
    )
    profile["require_missing_website"] = _parse_bool(
        payload.get("require_missing_website"), defaults["require_missing_website"]
    )
    profile["minimum_review_count"] = _parse_int(
        payload.get("minimum_review_count"), defaults["minimum_review_count"]
    )
    profile["minimum_rating"] = _parse_float(
        payload.get("minimum_rating"), defaults["minimum_rating"]
    )
    profile["scraper_depth"] = _parse_int(
        payload.get("scraper_depth"), defaults["scraper_depth"]
    )
    profile["scraper_concurrency"] = _parse_int(
        payload.get("scraper_concurrency"), defaults["scraper_concurrency"]
    )
    profile["scraper_radius"] = _parse_int(
        payload.get("scraper_radius"), defaults["scraper_radius"]
    )
    profile["scraper_fast_mode"] = _parse_bool(
        payload.get("scraper_fast_mode"), defaults["scraper_fast_mode"]
    )
    profile["scraper_lang"] = str(
        payload.get("scraper_lang", defaults["scraper_lang"])
    ).strip() or defaults["scraper_lang"]

    return profile


def _profile_summary(profile: dict[str, Any]) -> dict[str, Any]:
    target_queries = list(profile.get("target_queries", []))
    target_cities = list(profile.get("target_cities", []))
    target_niches = list(profile.get("target_niches", []))
    query_count = len(target_queries) or len(target_cities) * len(target_niches)
    return {
        "country": str(profile.get("country", get_country())),
        "target_cities": target_cities,
        "target_niches": target_niches,
        "target_queries": target_queries,
        "query_count_estimate": query_count,
        "must_have_phone": bool(profile.get("must_have_phone", get_must_have_phone())),
        "require_missing_website": bool(
            profile.get("require_missing_website", get_require_missing_website())
        ),
        "minimum_review_count": int(
            profile.get("minimum_review_count", get_minimum_review_count())
        ),
        "minimum_rating": float(profile.get("minimum_rating", get_minimum_rating())),
        "scraper_depth": int(profile.get("scraper_depth", get_scraper_depth())),
        "scraper_concurrency": int(
            profile.get("scraper_concurrency", get_scraper_concurrency())
        ),
        "scraper_radius": int(profile.get("scraper_radius", get_scraper_radius())),
        "scraper_fast_mode": bool(
            profile.get("scraper_fast_mode", get_scraper_fast_mode())
        ),
        "scraper_lang": str(profile.get("scraper_lang", get_scraper_lang())),
    }


def _scraper_output_files() -> list[Path]:
    base_path = Path(get_scraper_results_path())
    if not base_path.parent.exists():
        return []

    pattern = f"{base_path.stem}*{base_path.suffix}"
    return sorted(
        [item for item in base_path.parent.glob(pattern) if item.is_file()],
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    )


_PREPARED_RE = re.compile(
    r"Prepared (?P<queries>\d+) search queries(?: across (?P<batches>\d+) batch\(es\))?\."
)
_BATCH_START_RE = re.compile(
    r"Batch (?P<index>\d+)/(?P<total>\d+): scraping (?P<queries>\d+) query\(s\)\."
)
_BATCH_DONE_RE = re.compile(
    r"Batch (?P<index>\d+)/(?P<total>\d+): scanned (?P<rows>\d+) row\(s\), saved (?P<saved>\d+) lead\(s\) so far\."
)


def _mission_progress(
    logs: list[dict[str, str]],
    current_job: dict[str, Any] | None,
    job_history: list[dict[str, Any]],
) -> dict[str, Any]:
    job = current_job or (job_history[0] if job_history else None)
    if not job:
        return {}

    started_at = str(job.get("started_at", "")).strip()
    job_logs = [
        entry for entry in logs if not started_at or str(entry.get("timestamp", "")) >= started_at
    ]

    total_queries = int(job.get("meta", {}).get("query_count_estimate", 0) or 0)
    total_batches = 0
    batch_queries: dict[int, int] = {}
    batch_rows: dict[int, int] = {}
    batch_saved: dict[int, int] = {}
    latest_message = ""

    for entry in job_logs:
        message = str(entry.get("message", "")).strip()
        if not message:
            continue
        latest_message = message

        prepared_match = _PREPARED_RE.search(message)
        if prepared_match:
            total_queries = int(prepared_match.group("queries") or total_queries or 0)
            total_batches = int(prepared_match.group("batches") or total_batches or 1)

        start_match = _BATCH_START_RE.search(message)
        if start_match:
            index = int(start_match.group("index"))
            total_batches = max(total_batches, int(start_match.group("total")))
            batch_queries[index] = int(start_match.group("queries"))

        done_match = _BATCH_DONE_RE.search(message)
        if done_match:
            index = int(done_match.group("index"))
            total_batches = max(total_batches, int(done_match.group("total")))
            batch_rows[index] = int(done_match.group("rows"))
            batch_saved[index] = int(done_match.group("saved"))

    if total_batches <= 0:
        total_batches = 1 if total_queries else 0

    completed_batches = len(batch_saved)
    started_batches = len(batch_queries)
    running_batch = None
    for index in sorted(batch_queries):
        if index not in batch_saved:
            running_batch = index
            break

    if job.get("status") == "running" and running_batch is None and started_batches:
        running_batch = max(batch_queries)

    percent = 0
    if total_batches > 0:
        if job.get("status") == "completed":
            percent = 100
        elif job.get("status") in {"failed", "cancelled"}:
            percent = min(99, int((completed_batches / total_batches) * 100))
        else:
            percent = int((completed_batches / total_batches) * 100)
            if running_batch is not None:
                percent = min(95, percent + max(8, int(100 / (total_batches * 2))))

    phase = "Waiting to start."
    if latest_message:
        phase = latest_message
    if job.get("status") == "completed":
        phase = "Mission completed successfully."
    elif job.get("status") == "failed":
        phase = str(job.get("error", "")).strip() or "Mission failed."
    elif job.get("status") == "cancelled":
        phase = str(job.get("error", "")).strip() or "Mission was stopped."

    batches = []
    for index in range(1, total_batches + 1):
        status = "pending"
        if index in batch_saved:
            status = "completed"
        elif running_batch == index and job.get("status") == "running":
            status = "running"
        elif job.get("status") == "failed" and index in batch_queries:
            status = "failed"
        elif job.get("status") == "cancelled" and index in batch_queries:
            status = "cancelled"

        batches.append(
            {
                "index": index,
                "status": status,
                "queries": batch_queries.get(index, 0),
                "rows": batch_rows.get(index, 0),
                "saved": batch_saved.get(index, 0),
            }
        )

    return {
        "job_name": str(job.get("name", "")),
        "status": str(job.get("status", "")),
        "stop_requested": bool(job.get("stop_requested", False)),
        "started_at": started_at,
        "ended_at": str(job.get("ended_at", "")).strip(),
        "total_queries": total_queries,
        "total_batches": total_batches,
        "completed_batches": completed_batches,
        "running_batch": running_batch,
        "saved_so_far": max(batch_saved.values(), default=0),
        "rows_scanned": sum(batch_rows.values()),
        "percent": percent,
        "phase": phase,
        "batches": batches,
    }


class DashboardState:
    def __init__(self) -> None:
        self.lock = threading.Lock()
        self.logs: list[dict[str, str]] = []
        self.current_job: dict[str, Any] | None = None
        self.current_stop_event: threading.Event | None = None
        self.job_history: list[dict[str, Any]] = []
        self.drive_uploads: list[dict[str, Any]] = []
        self.last_discovery_profile = _profile_summary(_defaults_snapshot())
        self.max_logs = 400

    def log(self, level: str, message: str) -> None:
        entry = {"timestamp": _now(), "level": level, "message": message}
        with self.lock:
            self.logs.append(entry)
            self.logs = self.logs[-self.max_logs :]

    def set_last_discovery_profile(self, profile: dict[str, Any]) -> None:
        with self.lock:
            self.last_discovery_profile = _profile_summary(profile)

    def remember_drive_uploads(self, uploads: list[dict[str, Any]]) -> None:
        if not uploads:
            return

        with self.lock:
            self.drive_uploads = uploads + self.drive_uploads
            self.drive_uploads = self.drive_uploads[:12]

    def reset_view_state(self) -> None:
        with self.lock:
            self.logs = []
            self.job_history = []
            self.drive_uploads = []
            self.current_job = None
            self.current_stop_event = None

    def run_job(
        self,
        name: str,
        target: Callable[[threading.Event], dict[str, Any]],
        meta: dict[str, Any] | None = None,
    ) -> tuple[bool, dict[str, Any]]:
        with self.lock:
            if self.current_job and self.current_job.get("status") == "running":
                return False, dict(self.current_job)

            stop_event = threading.Event()
            job = {
                "id": str(uuid.uuid4())[:8],
                "name": name,
                "status": "running",
                "stop_requested": False,
                "meta": meta or {},
                "started_at": _now(),
                "ended_at": "",
                "result": None,
                "error": "",
            }
            self.current_job = job
            self.current_stop_event = stop_event

        def runner() -> None:
            self.log("job", f"{name} started")
            try:
                result = target(stop_event)
                job["status"] = "completed"
                job["result"] = result
                self.log("job", f"{name} completed")
            except AgentStoppedError as exc:
                job["status"] = "cancelled"
                job["error"] = str(exc)
                self.log("warning", f"{name} stopped")
            except Exception as exc:
                job["status"] = "failed"
                job["error"] = str(exc)
                self.log("error", f"{name} failed: {exc}")
            finally:
                job["ended_at"] = _now()
                with self.lock:
                    self.current_job = None
                    self.current_stop_event = None
                    self.job_history.insert(0, dict(job))
                    self.job_history = self.job_history[:12]

        threading.Thread(target=runner, daemon=True).start()
        return True, job

    def request_stop(self) -> tuple[bool, dict[str, Any] | None]:
        with self.lock:
            if not self.current_job or self.current_job.get("status") != "running":
                return False, None
            if self.current_stop_event is not None:
                self.current_stop_event.set()
            self.current_job["stop_requested"] = True
            job = dict(self.current_job)

        self.log("warning", f"Stop requested for {job['name']}.")
        return True, job

    def snapshot(self) -> dict[str, Any]:
        leads = sorted(
            load_leads(),
            key=lambda item: (item.get("score", 0), item.get("review_count", 0)),
            reverse=True,
        )
        status_counts = Counter(str(lead.get("status", "unknown")) for lead in leads)
        city_counts = Counter(str(lead.get("city", "Unknown")) for lead in leads)
        website_counts = Counter(
            str(lead.get("website_status", "unknown")) for lead in leads
        )

        with self.lock:
            current_job = dict(self.current_job) if self.current_job else None
            job_history = [dict(item) for item in self.job_history]
            logs = list(self.logs)
            drive_uploads = list(self.drive_uploads)
            last_discovery_profile = dict(self.last_discovery_profile)

        drive_status = DriveSync(logger=self.log).status()
        drive_status["recent_uploads"] = drive_uploads
        mission_progress = _mission_progress(logs, current_job, job_history)

        return {
            "generated_at": _now(),
            "metrics": {
                "total_leads": len(leads),
                "new_leads": status_counts.get("new", 0),
                "proposal_ready": status_counts.get("proposal_ready", 0),
                "won": status_counts.get("won", 0),
                "avg_score": round(
                    sum(float(lead.get("score", 0)) for lead in leads) / len(leads), 1
                )
                if leads
                else 0,
                "high_score": sum(1 for lead in leads if float(lead.get("score", 0)) >= 80),
                "missing_website": website_counts.get("missing", 0),
                "social_only": website_counts.get("social_only", 0),
                "marketplace_only": website_counts.get("marketplace_only", 0),
                "website_builder_only": website_counts.get("website_builder_only", 0),
            },
            "status_counts": dict(status_counts),
            "city_counts": city_counts.most_common(8),
            "website_counts": dict(website_counts),
            "leads": [
                {
                    "id": str(lead.get("id", "")),
                    "business_name": str(lead.get("business_name", "")),
                    "city": str(lead.get("city", "")),
                    "category": str(lead.get("category", "")),
                    "phone": str(lead.get("phone", "")),
                    "website_status": str(lead.get("website_status", "")),
                    "review_count": int(lead.get("review_count", 0)),
                    "review_rating": float(lead.get("review_rating", 0.0)),
                    "score": int(lead.get("score", 0)),
                    "status": str(lead.get("status", "")),
                    "proposal_path": str(lead.get("proposal_path", "")),
                    "build_package_path": str(lead.get("build_package_path", "")),
                    "maps_link": str(lead.get("maps_link", "")),
                    "notes": str(lead.get("notes", "")),
                    "duplicate_count": int(lead.get("duplicate_count", 1)),
                    "matched_queries": list(lead.get("matched_queries", [])),
                    "score_reasons": list(lead.get("score_reasons", [])),
                }
                for lead in leads[:60]
            ],
            "files": {
                "call_sheet": _file_meta(Path(get_call_sheet_path())),
                "leads_excel": _file_meta(Path(get_leads_excel_path())),
                "scraper_results": _file_meta(Path(get_scraper_results_path())),
                "scraper_batches": [_file_meta(path) for path in _scraper_output_files()[:6]],
                "proposals": _recent_files(Path(get_proposal_output_dir())),
                "build_packages": _recent_files(Path(get_build_package_output_dir())),
            },
            "job": current_job,
            "job_history": job_history,
            "mission_progress": mission_progress,
            "logs": logs,
            "lead_statuses": LEAD_STATUSES,
            "search_defaults": _defaults_snapshot(),
            "search_presets": _search_presets(),
            "last_discovery_profile": last_discovery_profile,
            "drive": drive_status,
        }


def _file_meta(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {"path": str(path), "exists": False}

    stat = path.stat()
    return {
        "path": str(path),
        "exists": True,
        "size": stat.st_size,
        "name": path.name,
        "modified_at": datetime.fromtimestamp(stat.st_mtime).strftime(
            "%Y-%m-%d %H:%M:%S"
        ),
    }


def _recent_files(directory: Path) -> list[dict[str, Any]]:
    if not directory.exists():
        return []

    files = sorted(
        [item for item in directory.iterdir() if item.is_file()],
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    )
    return [_file_meta(file) for file in files[:10]]


def _reset_runtime_outputs() -> dict[str, Any]:
    if state.current_job:
        raise RuntimeError("Stop the current job before resetting the workspace.")

    removed = 0
    for path in _scraper_output_files():
        if path.exists():
            path.unlink(missing_ok=True)
            removed += 1

    for single_path in [
        Path(get_call_sheet_path()),
        Path(get_scraper_results_path()),
        Path(get_leads_excel_path()),
    ]:
        if single_path.exists():
            single_path.unlink(missing_ok=True)
            removed += 1

    for directory in [Path(get_proposal_output_dir()), Path(get_build_package_output_dir())]:
        if directory.exists():
            for item in directory.iterdir():
                if item.is_file():
                    item.unlink(missing_ok=True)
                    removed += 1

    save_leads([])
    state.reset_view_state()
    state.log("info", "Workspace reset completed.")
    return {"removed_files": removed, "stored_leads": 0}


def _lead_rows() -> list[dict[str, Any]]:
    return sorted(
        load_leads(),
        key=lambda item: (item.get("score", 0), item.get("review_count", 0)),
        reverse=True,
    )


def _excel_workbook_bytes(leads: list[dict[str, Any]]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Leads"

    headers = [
        "Business",
        "City",
        "Category",
        "Phone",
        "Website Status",
        "Website",
        "Rating",
        "Review Count",
        "Score",
        "Status",
        "Why It Matters",
        "Matched Queries",
        "Google Maps Link",
        "Notes",
    ]
    sheet.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="12343B")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    for lead in leads:
        sheet.append(
            [
                str(lead.get("business_name", "")),
                str(lead.get("city", "")),
                str(lead.get("category", "")),
                str(lead.get("phone", "")),
                str(lead.get("website_status", "")),
                str(lead.get("website", "")),
                float(lead.get("review_rating", 0.0)),
                int(lead.get("review_count", 0)),
                int(lead.get("score", 0)),
                str(lead.get("status", "")),
                " | ".join(str(item) for item in lead.get("score_reasons", [])),
                " | ".join(str(item) for item in lead.get("matched_queries", [])),
                str(lead.get("maps_link", "")),
                str(lead.get("notes", "")),
            ]
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    widths = {
        "A": 30,
        "B": 18,
        "C": 18,
        "D": 18,
        "E": 18,
        "F": 28,
        "G": 10,
        "H": 12,
        "I": 10,
        "J": 18,
        "K": 44,
        "L": 34,
        "M": 36,
        "N": 28,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    payload = BytesIO()
    workbook.save(payload)
    payload.seek(0)

    with open(get_leads_excel_path(), "wb") as file:
        file.write(payload.getvalue())
    payload.seek(0)
    return payload


state = DashboardState()


def _agent(
    runtime_profile: dict[str, Any] | None = None,
    stop_event: threading.Event | None = None,
) -> WebsiteSalesAgent:
    return WebsiteSalesAgent(
        logger=state.log,
        runtime_profile=runtime_profile,
        stop_requested=stop_event.is_set if stop_event is not None else None,
    )


def _sync_paths(paths: list[str | Path]) -> list[dict[str, Any]]:
    unique_paths: list[Path] = []
    seen: set[str] = set()
    for item in paths:
        path = Path(item)
        if not path.exists():
            continue
        fingerprint = str(path.resolve())
        if fingerprint in seen:
            continue
        seen.add(fingerprint)
        unique_paths.append(path)

    drive = DriveSync(logger=state.log)
    drive_status = drive.status()
    if not drive_status.get("enabled"):
        return []

    if not drive_status.get("ready"):
        state.log("warning", str(drive_status.get("message", "Drive sync is not ready.")))
        return []

    uploads: list[dict[str, Any]] = []
    for path in unique_paths:
        try:
            uploads.append(drive.upload_file(path))
        except Exception as exc:
            state.log("warning", f"Drive sync skipped for {path.name}: {exc}")

    state.remember_drive_uploads(uploads)
    return uploads


def _discover(runtime_profile: dict[str, Any], stop_event: threading.Event) -> dict[str, Any]:
    state.set_last_discovery_profile(runtime_profile)
    leads = _agent(runtime_profile, stop_event=stop_event).discover_leads()
    uploads = _sync_paths([])
    return {
        "stored_leads": len(leads),
        "top_business": leads[0]["business_name"] if leads else "",
        "uploads": uploads,
    }


def _generate_briefs(
    stop_event: threading.Event, limit: int, status_filter: str
) -> dict[str, Any]:
    paths = _agent(stop_event=stop_event).generate_briefs(
        limit=limit, status_filter=status_filter
    )
    uploads = _sync_paths(paths)
    return {"generated_files": len(paths), "paths": paths, "uploads": uploads}


def _generate_builds(
    stop_event: threading.Event, limit: int, status_filter: str
) -> dict[str, Any]:
    paths = _agent(stop_event=stop_event).generate_build_packages(
        limit=limit, status_filter=status_filter
    )
    uploads = _sync_paths(paths)
    return {"generated_files": len(paths), "paths": paths, "uploads": uploads}


def _export_call_sheet(stop_event: threading.Event) -> dict[str, Any]:
    path = _agent(stop_event=stop_event).export_call_sheet()
    uploads = _sync_paths([path])
    return {"path": path, "uploads": uploads}


def _sync_exports(stop_event: threading.Event) -> dict[str, Any]:
    if stop_event.is_set():
        raise AgentStoppedError("Operation stopped by user.")

    paths: list[Path] = []
    call_sheet = Path(get_call_sheet_path())
    if call_sheet.exists():
        paths.append(call_sheet)

    paths.extend(Path(item["path"]) for item in _recent_files(Path(get_proposal_output_dir()))[:6])
    paths.extend(
        Path(item["path"]) for item in _recent_files(Path(get_build_package_output_dir()))[:6]
    )

    uploads = _sync_paths(paths)
    return {"uploaded": len(uploads), "uploads": uploads}


@app.get("/")
def dashboard() -> str:
    return render_template("dashboard.html")


@app.get("/api/state")
def api_state() -> Any:
    return jsonify(state.snapshot())


@app.post("/api/actions/discover")
def api_discover() -> Any:
    payload = request.get_json(silent=True) or {}
    profile = _build_discovery_profile(payload)
    started, job = state.run_job(
        "Lead discovery",
        lambda stop_event: _discover(profile, stop_event),
        _profile_summary(profile),
    )
    return jsonify({"started": started, "job": job}), 202 if started else 409


@app.post("/api/actions/briefs")
def api_briefs() -> Any:
    payload = request.get_json(silent=True) or {}
    limit = _parse_int(payload.get("limit"), 10)
    status_filter = str(payload.get("status_filter", "")).strip()
    started, job = state.run_job(
        "Generate proposal briefs",
        lambda stop_event: _generate_briefs(
            stop_event, limit=limit, status_filter=status_filter
        ),
        {"limit": limit, "status_filter": status_filter},
    )
    return jsonify({"started": started, "job": job}), 202 if started else 409


@app.post("/api/actions/build-packages")
def api_builds() -> Any:
    payload = request.get_json(silent=True) or {}
    limit = _parse_int(payload.get("limit"), 10)
    status_filter = str(payload.get("status_filter", "won")).strip() or "won"
    started, job = state.run_job(
        "Generate build packages",
        lambda stop_event: _generate_builds(
            stop_event, limit=limit, status_filter=status_filter
        ),
        {"limit": limit, "status_filter": status_filter},
    )
    return jsonify({"started": started, "job": job}), 202 if started else 409


@app.post("/api/actions/export-call-sheet")
def api_export() -> Any:
    started, job = state.run_job("Export call sheet", _export_call_sheet)
    return jsonify({"started": started, "job": job}), 202 if started else 409


@app.post("/api/actions/sync-exports")
def api_sync_exports() -> Any:
    started, job = state.run_job("Sync exports to Google Drive", _sync_exports)
    return jsonify({"started": started, "job": job}), 202 if started else 409


@app.post("/api/actions/stop")
def api_stop() -> Any:
    stopped, job = state.request_stop()
    if not stopped:
        return jsonify({"ok": False, "error": "No running job to stop."}), 409
    return jsonify({"ok": True, "job": job})


@app.post("/api/actions/reset")
def api_reset() -> Any:
    try:
        result = _reset_runtime_outputs()
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 409
    return jsonify({"ok": True, "result": result})


@app.get("/api/exports/leads.xlsx")
def api_export_leads_excel() -> Any:
    leads = _lead_rows()
    if not leads:
        return jsonify({"ok": False, "error": "No leads available to export."}), 404

    payload = _excel_workbook_bytes(leads)
    filename = f"website-leads-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
    return send_file(
        payload,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/api/leads/<lead_id>/status")
def api_update_status(lead_id: str) -> Any:
    payload = request.get_json(silent=True) or {}
    status_value = str(payload.get("status", "")).strip()
    notes = str(payload.get("notes", "")).strip()
    if status_value not in LEAD_STATUSES:
        return jsonify({"ok": False, "error": "Invalid status"}), 400

    updated = _agent().update_status(lead_id=lead_id, status=status_value, notes=notes)
    if updated is None:
        return jsonify({"ok": False, "error": "Lead not found"}), 404

    uploads: list[dict[str, Any]] = []
    build_package_path = str(updated.get("build_package_path", "")).strip()
    if build_package_path:
        uploads = _sync_paths([build_package_path])

    return jsonify({"ok": True, "lead": updated, "uploads": uploads})


def main() -> None:
    assert_folder_structure()

    render_port = os.environ.get("PORT", "").strip()
    host = os.environ.get(
        "MPRINTER_DASHBOARD_HOST", "0.0.0.0" if render_port else "127.0.0.1"
    )
    port = int(render_port or os.environ.get("MPRINTER_DASHBOARD_PORT", "5055"))
    browser_host = "127.0.0.1" if host == "0.0.0.0" else host
    url = f"http://{browser_host}:{port}"
    state.log("info", f"Dashboard ready at {url}")

    if render_port:
        state.log("info", "Running in deploy mode with public port binding enabled.")

    if os.environ.get("MPRINTER_OPEN_BROWSER", "1") == "1" and not render_port:
        threading.Timer(1.0, lambda: webbrowser.open(url)).start()

    app.run(host=host, port=port, debug=False, threaded=True)


if __name__ == "__main__":
    main()
